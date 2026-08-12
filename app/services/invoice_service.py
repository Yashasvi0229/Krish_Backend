"""
Invoice XLSX generator.

Produces an Excel invoice matching GNC's professional layout — headers,
client/claim/insured blocks, itemized service table, totals row.

Design notes:
    * Uses openpyxl (already a dep). No LibreOffice required at runtime.
    * Numbers are stored as native Excel numbers (not strings) so the
      reviewer can re-total in-cell if they edit.
    * Formulas are used for the subtotal / total rows — if the reviewer
      changes a line-item's rate or hours, the totals recompute. No
      "recalc.py" pipeline is needed because we never rely on cached
      values (see openpyxl gotcha in the skill: cached values are None
      after a Python write).
    * The layout intentionally uses only Excel-2007 functions to survive
      any later save/reopen cycle in LibreOffice.
"""
from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.config import settings
from app.utils.file_storage import storage


# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
FONT = "Arial"
THIN = Side(style="thin", color="666666")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADER_FONT = Font(name=FONT, size=18, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")

SUBHEADER_FONT = Font(name=FONT, size=11, bold=True)
SUBHEADER_FILL = PatternFill("solid", fgColor="D9E2F3")

BODY_FONT = Font(name=FONT, size=10)
BODY_BOLD = Font(name=FONT, size=10, bold=True)

TOTAL_FONT = Font(name=FONT, size=12, bold=True, color="FFFFFF")
TOTAL_FILL = PatternFill("solid", fgColor="1F4E79")


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------
def render_invoice_xlsx(
    *,
    invoice_no: str,
    invoice_date: date,
    gnc_file_no: str,
    client_details: dict[str, Any],
    insured_details: dict[str, Any],
    loss_details: dict[str, Any],
    line_items: list[dict[str, Any]],
    billing_period_start: date,
    billing_period_end: date,
    subtotal: Decimal,
    grand_total: Decimal,
    hourly_rate: Decimal,
    currency: str = "CAD",
) -> bytes:
    """Build the workbook in memory and return its bytes.

    `line_items` shape (each dict):
        {
            "line_number": 1,
            "description": "Review of RCV Estimate v2 (revised O&P)",
            "category": "Estimate",
            "rule_code": "RCV_ACV",
            "quantity_hours": 15.33,
            "rate": 150.00,
            "total": 2299.50,
            "source_email_id": "...",
            "source_attachment_id": "...",
            "ai_confidence": "High",
            "is_flagged": false,
            "reasoning": "...",
        }
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"

    # ---- Column widths ----
    for col, width in enumerate(
        [6, 42, 16, 12, 10, 14, 14], start=1
    ):
        ws.column_dimensions[get_column_letter(col)].width = width

    # ---- 1. Header ----
    ws.merge_cells("A1:G3")
    hdr = ws["A1"]
    hdr.value = "GNC GROUP INC.\nProperty Damage Consultants"
    hdr.font = HEADER_FONT
    hdr.fill = HEADER_FILL
    hdr.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 24
    ws.row_dimensions[3].height = 24

    # ---- 2. Invoice meta ----
    row = 5
    _labeled_pair(ws, row,   "A", "Invoice No.:",    invoice_no)
    _labeled_pair(ws, row,   "E", "Invoice Date:",   invoice_date.strftime("%d %b %Y"))
    _labeled_pair(ws, row+1, "A", "GNC File No.:",   gnc_file_no)
    _labeled_pair(ws, row+1, "E", "Billing Period:",
                  f"{billing_period_start.strftime('%d %b')} — "
                  f"{billing_period_end.strftime('%d %b %Y')}")

    # ---- 3. Client block ----
    row = 8
    _section_header(ws, row, "A", "G", "Client Details")
    row += 1
    _kv_block(ws, row, "A",
              [("Name:", client_details.get("name", "—")),
               ("Company:", client_details.get("company_legal_name", "—")),
               ("Email:", client_details.get("email", "—")),
               ("Phone:", client_details.get("phone", "—"))])

    # ---- 4. Insured block ----
    row = 14
    _section_header(ws, row, "A", "G", "Insured Details")
    row += 1
    _kv_block(ws, row, "A",
              [("Insured Name:", insured_details.get("insured_name", "—")),
               ("Claim No.:", loss_details.get("claim_no", "—")),
               ("File Name:", loss_details.get("file_name", "—")),
               ("Loss Type:", loss_details.get("loss_type", "Unknown"))])

    # ---- 5. Line items table ----
    row = 20
    _section_header(ws, row, "A", "G", "Services Rendered")
    row += 1

    # Column headers
    headers = ["#", "Description", "Category", "Rule", "Hours", "Rate", "Amount"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = SUBHEADER_FONT
        c.fill = SUBHEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER_ALL
    ws.row_dimensions[row].height = 20

    # Data rows
    row += 1
    start_data_row = row
    for item in line_items:
        _line_item_row(
            ws, row,
            line_no=item.get("line_number", row - start_data_row + 1),
            description=item.get("description", ""),
            category=item.get("category", ""),
            rule_code=item.get("rule_code", ""),
            hours=float(item.get("quantity_hours") or 0),
            rate=float(item.get("rate") or hourly_rate),
        )
        row += 1

    end_data_row = row - 1
    if end_data_row < start_data_row:
        # No lines — add a placeholder row so the table looks valid.
        ws.cell(row=start_data_row, column=2,
                value="(No billable items — reviewer to add manually)"
                ).font = BODY_FONT
        end_data_row = start_data_row
        row = start_data_row + 1

    # ---- 6. Totals row ----
    row += 1
    total_label_cell = ws.cell(row=row, column=6, value="TOTAL")
    total_label_cell.font = TOTAL_FONT
    total_label_cell.fill = TOTAL_FILL
    total_label_cell.alignment = Alignment(horizontal="right", vertical="center")
    total_label_cell.border = BORDER_ALL

    total_val_cell = ws.cell(
        row=row, column=7,
        value=f"=SUM(G{start_data_row}:G{end_data_row})",
    )
    total_val_cell.font = TOTAL_FONT
    total_val_cell.fill = TOTAL_FILL
    total_val_cell.alignment = Alignment(horizontal="right", vertical="center")
    total_val_cell.border = BORDER_ALL
    total_val_cell.number_format = f'"{_currency_symbol(currency)}"#,##0.00'
    ws.row_dimensions[row].height = 22

    # ---- 7. Footer ----
    row += 3
    _kv_block(ws, row, "A",
              [("Hourly Rate:", f"{_currency_symbol(currency)}{hourly_rate:.2f} / hr"),
               ("Currency:", currency),
               ("Generated At:", datetime.now().strftime("%d %b %Y, %H:%M"))])

    # ---- Serialize ----
    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _labeled_pair(ws, row: int, col: str, label: str, value: str) -> None:
    l = ws[f"{col}{row}"]
    l.value = label
    l.font = BODY_BOLD
    # value cell is next column
    from openpyxl.utils.cell import column_index_from_string
    v_col = column_index_from_string(col) + 1
    v = ws.cell(row=row, column=v_col, value=str(value))
    v.font = BODY_FONT


def _section_header(ws, row: int, col_from: str, col_to: str, text: str) -> None:
    ws.merge_cells(f"{col_from}{row}:{col_to}{row}")
    c = ws[f"{col_from}{row}"]
    c.value = text
    c.font = SUBHEADER_FONT
    c.fill = SUBHEADER_FILL
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 18


def _kv_block(ws, row: int, col: str, pairs: list[tuple[str, str]]) -> None:
    from openpyxl.utils.cell import column_index_from_string
    base = column_index_from_string(col)
    for i, (k, v) in enumerate(pairs):
        r = row + i
        lc = ws.cell(row=r, column=base, value=k)
        lc.font = BODY_BOLD
        vc = ws.cell(row=r, column=base + 1, value=str(v))
        vc.font = BODY_FONT


def _line_item_row(
    ws, row: int,
    *, line_no: int, description: str, category: str, rule_code: str,
    hours: float, rate: float,
) -> None:
    """Write one billable line row. Uses formulas for the amount column
    so reviewer edits recompute in-cell."""
    cells = [
        (1, line_no,                   "center"),
        (2, description,               "left"),
        (3, category,                  "left"),
        (4, rule_code,                 "center"),
        (5, hours,                     "right"),
        (6, rate,                      "right"),
        (7, f"=E{row}*F{row}",         "right"),
    ]
    for col, val, align in cells:
        c = ws.cell(row=row, column=col, value=val)
        c.font = BODY_FONT
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=(col == 2))
        c.border = BORDER_ALL

    ws.cell(row=row, column=5).number_format = '0.00'
    ws.cell(row=row, column=6).number_format = '"$"#,##0.00'
    ws.cell(row=row, column=7).number_format = '"$"#,##0.00'
    ws.row_dimensions[row].height = 24


def _currency_symbol(currency: str) -> str:
    return {"USD": "$", "CAD": "C$", "INR": "₹", "EUR": "€", "GBP": "£"}.get(currency, "$")


# ---------------------------------------------------------------------------
# Save to disk
# ---------------------------------------------------------------------------
def save_invoice_file(invoice_no: str, gnc_file_no: str, data: bytes) -> tuple[str, int]:
    """Persist to the storage root under invoices/{gnc_file_no}/{invoice_no}.xlsx.
    Returns (relative_path, size_bytes)."""
    safe_gnc = "".join(c for c in gnc_file_no if c.isalnum() or c in "-_") or "unknown"
    rel_path = f"invoices/{safe_gnc}/{invoice_no}.xlsx"
    storage.write_bytes(rel_path, data)
    return rel_path, len(data)
